from typing import Any
import uuid
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Protection

from app.utils import check_columns_exist


def create_workbook(name: str, sheet_name: list, type: str):
    wb = Workbook()
    wbs = []
    for index, sheet in enumerate(sheet_name):
        wbs.append(sheet)
        wbs[index] = wb.create_sheet(sheet, index)
        wbs[index].title = sheet
    wb.save(filename=f'files/excel/{type}/{name}.xlsx')


def write_data_title(name: str, sheet_name: str, columns: list, type: str):
    wb = load_workbook(f'files/excel/{type}/{name}.xlsx')
    sheet = wb.get_sheet_by_name(sheet_name)
    gris = "00C0C0C0"
    for index, colum in enumerate(columns):
        sheet.cell(row=1, column=index + 1).value = colum
        sheet.cell(row=1, column=index + 1).fill = PatternFill(start_color=gris, end_color=gris, fill_type="solid")

    wb.save(filename=f'files/excel/{type}/{name}.xlsx')


def insert_data_xlsx(name: str, sheet_name: str, all_data: Any, columns: list, type: str):
    wb = load_workbook(f'files/excel/{type}/{name}.xlsx')
    sheet = wb.get_sheet_by_name(sheet_name)
    sheet.protection.sheet = True
    row = 2

    for index_, data in enumerate(all_data):
        for index, col in enumerate(data):
            sheet.cell(row=row, column=index + 1).value = str(data[index])
        row += 1

    gris = "00C0C0C0"
    row = 2
    for index_, data in enumerate(all_data):
        for index, col in enumerate(data):
            value_ = str(sheet.cell(row=1, column=index + 1).value)
            if value_[0:3] == "ec_":
                sheet.cell(row=row, column=index + 1).protection = Protection(locked=False, hidden=False)
            if value_[0:3] != "ec_":
                sheet.cell(row=row, column=index + 1).fill = PatternFill(start_color=gris,
                                                                         end_color=gris, fill_type="solid")
        row += 1

    wb.save(filename=f'files/excel/{type}/{name}.xlsx')
    return f'files/excel/{type}/{name}.xlsx'


def insert_from_xlsx(name: str, sheet_name: str, all_data: Any, columns: list, type: str):
    wb = load_workbook(f'files/excel/{type}/{name}.xlsx')


def get_all_sheet(workbook: str):
    wb = load_workbook(workbook)
    return wb.sheetnames


def validation_file(name: str, sheet_name: str, schemas: str) -> str:
    wb = load_workbook(name)
    sheet = wb.get_sheet_by_name(sheet_name)
    columns = check_columns_exist(schemas, sheet_name)
    print(columns)
    for col in range(sheet.max_column):
        if str(sheet.cell(row=1, column=col + 1).value) != columns[col]:
            return f"invalid columns {columns[col]} and {sheet.cell(row=1, column=col + 1).value} is differents"
    return "valid"


def validation_file_note(name: str, sheet_name: str, table_name: str, schemas: str) -> str:
    wb = load_workbook(name)
    sheet = wb.get_sheet_by_name(sheet_name)
    columns = check_columns_exist(schemas, table_name)
    for col in range(sheet.max_column-1):
        if str(sheet.cell(row=1, column=col + 1).value) != columns[col]:
            return f"invalid columns {columns[col]} and {sheet.cell(row=1, column=col + 1).value} is differents"
    return "valid"


def get_data_xlsx(name: str, sheet_name: str) -> Any:
    wb = load_workbook(name)
    sheet = wb.get_sheet_by_name(sheet_name)
    all_data = []
    for row in range(sheet.max_row):
        data = {}
        if row != 0:
            data['uuid'] = str(uuid.uuid4())
            for col in range(sheet.max_column - 1):
                data[str(sheet.cell(row=1, column=col + 2).value)] = str(sheet.cell(row=row + 1, column=col + 2).value)
            all_data.append(data)
    return all_data


def get_data_xlsx_note(name: str, sheet_name: str) -> Any:
    wb = load_workbook(name)
    sheet = wb.get_sheet_by_name(sheet_name)
    all_data = []
    for row in range(sheet.max_row):
        data = {}
        if row != 0:
            for col in range(sheet.max_column):
                value_note = str(sheet.cell(row=row + 1, column=col + 1).value)
                num_carte = str(sheet.cell(row=row + 1, column=1).value)
                if value_note == "None":
                    value_note = None
                if num_carte != "None":
                    data[str(sheet.cell(row=1, column=col + 1).value)] = value_note
            if len(data) != 0:
                all_data.append(data)
    return all_data
